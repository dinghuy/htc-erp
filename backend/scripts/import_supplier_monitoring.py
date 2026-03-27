import argparse
import json
import re
import sqlite3
import sys
import uuid
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


EXCEL_ERRORS = {"#VALUE!", "#REF!", "#DIV/0!", "#N/A", "#NAME?", "#NULL!"}


def normalize_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", (value or "").strip().lower())


def clean_text(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text or text.upper() in EXCEL_ERRORS:
        return ""
    return re.sub(r"\s+", " ", text)


def normalize_tag_string(raw: str) -> str:
    values = re.split(r"[,;\n|]+", str(raw or ""))
    seen = set()
    tags = []
    for value in values:
        tag = clean_text(value)
        if not tag:
            continue
        key = tag.lower()
        if key in seen:
            continue
        seen.add(key)
        tags.append(tag)
    return ", ".join(tags)


def split_name(full_name: str) -> tuple[str, str]:
    text = clean_text(full_name)
    if not text:
        return "", ""
    if "," in text:
        left, right = [part.strip() for part in text.split(",", 1)]
        return right, left
    parts = text.split(" ")
    if len(parts) == 1:
        return "", parts[0]
    return " ".join(parts[:-1]), parts[-1]


def map_status(raw: str) -> str:
    text = normalize_key(raw)
    if not text:
        return "active"
    if text in {"active", "approved", "current", "open", "new"}:
        return "active"
    if text in {"inactive", "closed", "archived", "disabled"}:
        return "inactive"
    return raw.strip().lower() if raw.strip() else "active"


def build_description(description: str, note: str) -> str:
    parts = [clean_text(description)]
    note_value = clean_text(note)
    if note_value:
        parts.append(f"Note: {note_value}")
    return "\n".join(part for part in parts if part)


def fetch_one(cur: sqlite3.Cursor, sql: str, params=()):
    cur.execute(sql, params)
    return cur.fetchone()


def create_backup(db_path: Path) -> Path:
    backup_path = db_path.with_name(f"{db_path.stem}.backup-{datetime.now().strftime('%Y%m%d-%H%M%S')}{db_path.suffix}")
    source = sqlite3.connect(db_path)
    target = sqlite3.connect(backup_path)
    try:
        source.backup(target)
    finally:
        target.close()
        source.close()
    return backup_path


def load_suppliers(ws):
    suppliers = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = clean_text(row[4] if len(row) > 4 else "")
        company = clean_text(row[5] if len(row) > 5 else "")
        description = clean_text(row[6] if len(row) > 6 else "")
        tag = clean_text(row[7] if len(row) > 7 else "")
        country = clean_text(row[9] if len(row) > 9 else "")
        website = clean_text(row[10] if len(row) > 10 else "")
        address = clean_text(row[11] if len(row) > 11 else "")
        status = clean_text(row[13] if len(row) > 13 else "")
        note = clean_text(row[14] if len(row) > 14 else "")

        if not code and not company:
            continue

        supplier = {
            "code": code,
            "companyName": company or code,
            "shortName": code.split("-", 1)[0].strip() if "-" in code else "",
            "description": build_description(description, note),
            "tag": normalize_tag_string(tag),
            "industry": normalize_tag_string(tag),
            "country": country,
            "website": website,
            "address": address,
            "status": map_status(status),
        }
        suppliers.append(supplier)
    return suppliers


def load_contacts(ws):
    contacts = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = clean_text(row[1] if len(row) > 1 else "")
        nickname = clean_text(row[2] if len(row) > 2 else "")
        department = clean_text(row[3] if len(row) > 3 else "")
        title = clean_text(row[4] if len(row) > 4 else "")
        company_ref = clean_text(row[5] if len(row) > 5 else "")
        email_1 = clean_text(row[8] if len(row) > 8 else "")
        email_2 = clean_text(row[9] if len(row) > 9 else "")
        mobile = clean_text(row[10] if len(row) > 10 else "")
        phone = clean_text(row[11] if len(row) > 11 else "")

        if not company_ref or not (name or nickname or email_1 or email_2 or mobile or phone):
            continue

        full_name = name or nickname
        first_name, last_name = split_name(full_name)
        contacts.append(
            {
                "fullName": full_name,
                "firstName": first_name,
                "lastName": last_name,
                "department": department,
                "jobTitle": title,
                "companyRef": company_ref,
                "email": email_1 or email_2,
                "phone": mobile or phone,
            }
        )
    return contacts


def resolve_supplier_id(cur: sqlite3.Cursor, supplier: dict):
    code = supplier["code"]
    company_name = supplier["companyName"]
    short_name = supplier["shortName"]

    if code:
        row = fetch_one(
            cur,
            "SELECT id FROM Account WHERE accountType='Supplier' AND LOWER(TRIM(code)) = LOWER(TRIM(?)) LIMIT 1",
            (code,),
        )
        if row:
            return row[0]

    row = fetch_one(
        cur,
        "SELECT id FROM Account WHERE accountType='Supplier' AND LOWER(TRIM(companyName)) = LOWER(TRIM(?)) LIMIT 1",
        (company_name,),
    )
    if row:
        return row[0]

    if short_name:
        row = fetch_one(
            cur,
            "SELECT id FROM Account WHERE accountType='Supplier' AND LOWER(TRIM(shortName)) = LOWER(TRIM(?)) LIMIT 1",
            (short_name,),
        )
        if row:
            return row[0]

    return None


def upsert_suppliers(cur: sqlite3.Cursor, suppliers: list[dict], summary: dict):
    alias_map = {}

    for supplier in suppliers:
        supplier_id = resolve_supplier_id(cur, supplier)
        if supplier_id:
            cur.execute(
                """
                UPDATE Account
                SET companyName = ?,
                    industry = ?,
                    website = ?,
                    address = ?,
                    status = ?,
                    accountType = 'Supplier',
                    code = ?,
                    shortName = ?,
                    description = ?,
                    tag = ?,
                    country = ?
                WHERE id = ?
                """,
                (
                    supplier["companyName"],
                    supplier["industry"],
                    supplier["website"],
                    supplier["address"],
                    supplier["status"],
                    supplier["code"],
                    supplier["shortName"],
                    supplier["description"],
                    supplier["tag"],
                    supplier["country"],
                    supplier_id,
                ),
            )
            summary["suppliersUpdated"] += 1
        else:
            supplier_id = str(uuid.uuid4())
            cur.execute(
                """
                INSERT INTO Account (
                    id, companyName, industry, website, address, status,
                    accountType, code, shortName, description, tag, country
                ) VALUES (?, ?, ?, ?, ?, ?, 'Supplier', ?, ?, ?, ?, ?)
                """,
                (
                    supplier_id,
                    supplier["companyName"],
                    supplier["industry"],
                    supplier["website"],
                    supplier["address"],
                    supplier["status"],
                    supplier["code"],
                    supplier["shortName"],
                    supplier["description"],
                    supplier["tag"],
                    supplier["country"],
                ),
            )
            summary["suppliersInserted"] += 1

        for alias in {supplier["code"], supplier["companyName"], supplier["shortName"]}:
            alias_key = normalize_key(alias)
            if alias_key:
                alias_map[alias_key] = supplier_id

    return alias_map


def resolve_contact_account_id(cur: sqlite3.Cursor, company_ref: str, alias_map: dict[str, str]):
    alias_key = normalize_key(company_ref)
    if alias_key in alias_map:
        return alias_map[alias_key]

    row = fetch_one(
        cur,
        """
        SELECT id
        FROM Account
        WHERE accountType='Supplier'
          AND (
            LOWER(TRIM(code)) = LOWER(TRIM(?))
            OR LOWER(TRIM(companyName)) = LOWER(TRIM(?))
            OR LOWER(TRIM(shortName)) = LOWER(TRIM(?))
          )
        LIMIT 1
        """,
        (company_ref, company_ref, company_ref),
    )
    return row[0] if row else None


def upsert_contacts(cur: sqlite3.Cursor, contacts: list[dict], alias_map: dict[str, str], summary: dict):
    primary_contact_seen = defaultdict(bool)

    for contact in contacts:
        account_id = resolve_contact_account_id(cur, contact["companyRef"], alias_map)
        if not account_id:
            summary["contactsSkippedUnmatchedSupplier"] += 1
            continue

        row = None
        if contact["email"]:
            row = fetch_one(
                cur,
                "SELECT id FROM Contact WHERE accountId = ? AND LOWER(TRIM(email)) = LOWER(TRIM(?)) LIMIT 1",
                (account_id, contact["email"]),
            )

        if not row and contact["fullName"]:
            row = fetch_one(
                cur,
                """
                SELECT id
                FROM Contact
                WHERE accountId = ?
                  AND LOWER(TRIM(COALESCE(firstName, '') || ' ' || COALESCE(lastName, ''))) = LOWER(TRIM(?))
                LIMIT 1
                """,
                (account_id, contact["fullName"]),
            )

        is_primary = 0 if primary_contact_seen[account_id] else 1
        primary_contact_seen[account_id] = True

        if row:
            cur.execute(
                """
                UPDATE Contact
                SET lastName = ?,
                    firstName = ?,
                    department = ?,
                    jobTitle = ?,
                    email = ?,
                    phone = ?,
                    isPrimaryContact = CASE WHEN isPrimaryContact = 1 THEN 1 ELSE ? END
                WHERE id = ?
                """,
                (
                    contact["lastName"],
                    contact["firstName"],
                    contact["department"],
                    contact["jobTitle"],
                    contact["email"],
                    contact["phone"],
                    is_primary,
                    row[0],
                ),
            )
            summary["contactsUpdated"] += 1
        else:
            cur.execute(
                """
                INSERT INTO Contact (
                    id, accountId, lastName, firstName, department, jobTitle,
                    email, phone, isPrimaryContact
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    str(uuid.uuid4()),
                    account_id,
                    contact["lastName"],
                    contact["firstName"],
                    contact["department"],
                    contact["jobTitle"],
                    contact["email"],
                    contact["phone"],
                    is_primary,
                ),
            )
            summary["contactsInserted"] += 1


def main():
    parser = argparse.ArgumentParser(description="Import SupplierMonitoring XLSM into CRM supplier tables.")
    parser.add_argument("--file", required=True, help="Path to SupplierMonitoring2025.xlsm")
    parser.add_argument("--db", default=str(Path(__file__).resolve().parents[1] / "crm.db"), help="Path to SQLite DB")
    parser.add_argument("--dry-run", action="store_true", help="Parse and simulate without committing")
    args = parser.parse_args()

    workbook_path = Path(args.file).resolve()
    db_path = Path(args.db).resolve()
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")
    if not db_path.exists():
        raise FileNotFoundError(f"Database not found: {db_path}")

    wb = load_workbook(workbook_path, read_only=True, data_only=True, keep_vba=True)
    supplier_sheet = wb["SupplierList"]
    contacts_sheet = wb["Contacts"]

    suppliers = load_suppliers(supplier_sheet)
    contacts = load_contacts(contacts_sheet)

    summary = {
        "dryRun": args.dry_run,
        "workbook": str(workbook_path),
        "database": str(db_path),
        "backupPath": None,
        "supplierRowsParsed": len(suppliers),
        "contactRowsParsed": len(contacts),
        "suppliersInserted": 0,
        "suppliersUpdated": 0,
        "contactsInserted": 0,
        "contactsUpdated": 0,
        "contactsSkippedUnmatchedSupplier": 0,
    }

    if not args.dry_run:
        summary["backupPath"] = str(create_backup(db_path))

    con = sqlite3.connect(db_path)
    try:
        cur = con.cursor()
        cur.execute("BEGIN")
        alias_map = upsert_suppliers(cur, suppliers, summary)
        upsert_contacts(cur, contacts, alias_map, summary)

        if args.dry_run:
            con.rollback()
        else:
            con.commit()
    except Exception:
        con.rollback()
        raise
    finally:
        con.close()

    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(json.dumps({"error": str(exc)}), file=sys.stderr)
        sys.exit(1)
